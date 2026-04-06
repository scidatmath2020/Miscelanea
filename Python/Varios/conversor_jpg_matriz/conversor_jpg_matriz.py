# -*- coding: utf-8 -*-
"""
Created on Mon Apr 6 00:52:29 2026
@author: SciData
"""

'''
Indicas el nombre de tu imagen. Debe estar en formato jpg
Indicas la carpeta donde se encuentra tu imagen

Devuelve
- La matriz csv de la imagen en blanco y negro
- El excel con la imagen coloreada en escala de grises
'''

# pip install openpyxl
import numpy as np
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

#%%

###############################################################################
###############################################################################
###############################################################################

# ==================== CONFIGURACIÓN ====================

nombre_archivo = '...........jpg'
ruta = r'..........................'

###############################################################################
###############################################################################
###############################################################################

# ==================== PROCESAMIENTO DE IMAGEN ====================
img = Image.open(ruta + f'\{nombre_archivo}').convert('L')  # Blanco y negro
matriz_grises = np.array(img).astype(np.float32)
matriz_normalizada = matriz_grises / 255.0

print(f"Forma de la matriz: {matriz_normalizada.shape}")
print(f"Valores entre {matriz_normalizada.min():.4f} y {matriz_normalizada.max():.4f}")

# ==================== CREAR NOMBRES DE COLUMNAS ====================
num_columnas = matriz_normalizada.shape[1]
nombres_columnas = [f'col{i+1}' for i in range(num_columnas)]

# ==================== GUARDAR CSV CON NOMBRES DE COLUMNAS ====================
df = pd.DataFrame(matriz_normalizada, columns=nombres_columnas)

csv_path = ruta + f'\{nombre_archivo[:-4]}_gray.csv'
df.to_csv(csv_path, index=False, float_format='%.6f')   # header=True por defecto
print(f"✅ CSV guardado en: {csv_path}")

# ==================== GUARDAR EXCEL CON FORMATO ESPECIAL ====================
excel_path = ruta + f'\{nombre_archivo[:-4]}_gray.xlsx'

# Guardar primero con pandas (incluyendo los nombres de columnas)
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Sheet1', index=False, float_format='%.6f')

# Cargar el archivo con openpyxl para aplicar formatos finos
wb = load_workbook(excel_path)
ws = wb.active

# ====================== ANCHO DE COLUMNAS ======================
ancho_columna = 0.08
for col in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].width = ancho_columna

# ====================== ALTO DE FILAS ======================
alto_fila = 0.75
for row in range(1, ws.max_row + 1):
    ws.row_dimensions[row].height = alto_fila

# ====================== FORMATO CONDICIONAL 2 COLORES ======================
# Negro (valor mínimo) → Blanco (valor máximo)
color_scale_rule = ColorScaleRule(
    start_type='min',
    start_color='000000',   # Negro
    end_type='max',
    end_color='FFFFFF'      # Blanco
)

# Aplicar a toda la matriz (incluyendo la fila de encabezados)
rango = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
ws.conditional_formatting.add(rango, color_scale_rule)

# Guardar el Excel con todos los cambios
wb.save(excel_path)

print(f"✅ Excel guardado correctamente con formato especial en:")
print(excel_path)
print(f"Tamaño: {matriz_normalizada.shape[0]} filas × {matriz_normalizada.shape[1]} columnas")
